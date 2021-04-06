"""
批量生成测试数据脚本
思路：大数据量的时候，先把数据存入列表或者字典，统一写入excel

"""
import openpyxl
from faker import Faker
from pypinyin import lazy_pinyin,Style
import random


#生成name
f = Faker(locale='zh_CN')
name_list = []
for _ in range(50000):
    name_list.append(f.name())

# print(name_list)


#写入测试数据到excel文件中
wb = openpyxl.Workbook()   #自动生成文件
# wb = openpyxl.load_workbook(filename='data.xlsx')   #读取已存在的文件
ws = wb.active

#写入首行
header_list = ['tts_type','tts_content','tts_content_pinyin','tts_name','tts_savepath']
ws.append(header_list)

#写入测试数据
for i in range(len(name_list)): # 第几行-逐行写入数据
    # 汉字转换成拼音
    pinyin_name = lazy_pinyin(hans=name_list[i],style=Style.NORMAL)
    # print(pinyin_name)
    tts_content_pinyin = ''
    for n in pinyin_name:
        tts_content_pinyin +=' ' + ''.join(n+random.choice('1234'))

    # print(tts_content_pinyin)

    ws.cell(row=i+2,column=1,value='name')
    ws.cell(row=i+2,column=2,value=name_list[i])
    ws.cell(row=i+2,column=3,value=tts_content_pinyin)     #将汉字转换成拼音
    ws.cell(row=i+2,column=4,value='12358864_1_'+name_list[i]+'_name.wav')
    ws.cell(row=i+2,column=5,value='/Users/admin/PycharmProjects/ttsplots')

#保存文件
wb.save(filename='data.xlsx')
print('ok')
