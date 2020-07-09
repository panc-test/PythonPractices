'''
写入excel文件
对象：
（1）workbook
（2）sheet
（3）cell
'''

from openpyxl import Workbook

#创建工作簿
wb=Workbook()
#调用创建工作簿的时候自动新建的sheet工作表
sheet=wb.active

#创建工作表
sheet1=wb.create_sheet()
sheet2=wb.create_sheet(title='xxx')

#写入数据
sheet['A1']=0
sheet1['A1']=100
sheet2['A1']=200

#重命名工作表
sheet1.title='test'
print(wb.sheetnames)
#保存工作表，会覆盖掉源文件中的数据
wb.save('./data.xlsx')






