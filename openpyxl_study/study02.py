'''
读取excel文件的测试数据，转化成字典格式存储
'''
from openpyxl import load_workbook

#读取excel文件
wb=load_workbook(filename='file.xlsx')
ws=wb.active

#读取头部
for cells in ws.iter_rows(min_col=1,max_row=1,values_only=True):
    print('headers=',cells)
    headers=cells

#定义一个空的字典，用来存放读取到的所有测试数据
file_data={}

#读取测试数据
for cell in ws.iter_rows(min_col=1,min_row=2,values_only=True):
    print('测试数据',cell)
    #将读取的测试数据按照键值对的格式id:data，存入字典中
    data_id=cell[0]
    #定义一个空的字典，用来存放读取到的每一行测试数据
    data={}
    data['username']=cell[1]
    data['password']=cell[2]
    data['except_val']=cell[3]
    file_data[data_id]=data

print('从excel文件中提取的测试数据=',file_data)





