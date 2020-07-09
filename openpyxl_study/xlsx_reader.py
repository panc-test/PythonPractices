'''
读取excel文件:只能读取.xlsx文件，不能读取.xls文件
'''

from openpyxl import load_workbook

#读取工作簿
wb=load_workbook(filename='data.xlsx')

#读取工作表
sheet=wb['test']

#读取单元格
A1=sheet.cell(row=1,column=1).value

print(A1)



