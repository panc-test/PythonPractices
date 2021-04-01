"""
cell 单元格读写操作

"""
import openpyxl

wb = openpyxl.load_workbook('data.xlsx')

ws = wb.active
#查询
print(ws.cell(row=4, column=2).value)
print(ws['A4'].value)
#
# #插入数据
# ws.cell(row=3,column=6,value=36)
#
# #保存文件
# wb.save(filename='data.xlsx')
