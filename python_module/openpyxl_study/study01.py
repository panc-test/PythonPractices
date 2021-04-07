'''
openpyxl学习地址：https://www.yuque.com/ifd39g/zdmqip/revq81#a-simple-approach-to-reading-an-excel-spreadsheet

'''

#Reading an Excel Spreadsheet
from openpyxl import load_workbook

#工作簿
workbook = load_workbook(filename="sample.xlsx")
print(workbook.sheetnames)

#工作表
worksheet=workbook.active
print(worksheet.title)
print('表格的最大列=',worksheet.max_column,'表格的最大行=',worksheet.max_row,'表格的最小列=',worksheet.min_column,'表格的最小行=',worksheet.min_row)

rows=worksheet.iter_cols(min_col=1,max_col=3,min_row=7,max_row=10,values_only=True)
for i in rows:
    print(i)

#单元格
cell=worksheet.cell(row=7,column=3)
print('第7行3列单元格=',cell.value)

