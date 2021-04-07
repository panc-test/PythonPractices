'''
读取excel文件:只能读取.xlsx文件，不能读取.xls文件
'''

from openpyxl import load_workbook,worksheet

#读取工作簿
wb=load_workbook(filename='data.xlsx')

#读取工作表
ws=wb['test']

#读取单元格
A1=ws.cell(row=1,column=1).value
print(A1)

#按行读取
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
    print(row)

#按列读取
for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    print(col)

# iterate through all the rows or columns
cells_row=tuple(ws.rows)
cells_col=tuple(ws.columns)
print(cells_row)
print(cells_col)


